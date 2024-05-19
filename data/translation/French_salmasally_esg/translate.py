import time
import pandas as pd
import requests
import re
import unicodedata
from googletrans import Translator
from openpyxl import load_workbook, Workbook


def parquet_to_excel(parquet_path, excel_path):
    df = pd.read_parquet(parquet_path, engine='pyarrow')
    df.to_excel(excel_path, index=False)

    print(f"Data has been successfully written to {excel_path}")

def deepl_translate_excel(input_excel_path, output_excel_path, target_language, deepl_api_key):
    df = pd.read_excel(input_excel_path)

    # DeepL API URL
    url = "https://api-free.deepl.com/v2/translate"

    for column in df.columns:
        for i in range(len(df)):
            text = df.at[i, column]
            if pd.notna(text):  # None
                params = {
                    'auth_key': deepl_api_key,
                    'text': text,
                    'source_lang': 'KO',
                    'target_lang': target_language
                }
                response = requests.post(url, data=params)
                if response.status_code == 200:
                    translated_text = response.json()['translations'][0]['text']
                    df.at[i, column] = translated_text
                else:
                    print(f"Error translating text: {response.text}")

    # new excel 
    df.to_excel(output_excel_path, index=False)

    print(f"Translation completed and saved to {output_excel_path}")

def convert_csv_to_xlsx(csv_file_path, xlsx_file_path):
    # Read the CSV file
    data = pd.read_csv(csv_file_path)
    # Save the data to an Excel file
    data.to_excel(xlsx_file_path, index=False)

def is_hangul(text):
    # check korean
    return any('\uac00' <= char <= '\ud7a3' for char in text)

def is_french(text):
    # check French
    french_accents = {'À', 'Â', 'Æ', 'Ç', 'É', 'È', 'Ê', 'Ë', 'Î', 'Ï', 'Ô', 'Œ', 'Ù', 'Û', 'Ü', 'Ÿ', 'à', 'â', 'æ', 'ç', 'é', 'è', 'ê', 'ë', 'î', 'ï', 'ô', 'œ', 'ù', 'û', 'ü', 'ÿ'}
    return any(char in french_accents or unicodedata.name(char).startswith('LATIN') for char in text)

def contains_html(text):
    # check html
    return bool(re.search(r'<[^>]+>', text))

def google_translate_excel(input_file_path, output_file_path, sheet_index):
    translator = Translator()

    workbook = load_workbook(filename=input_file_path)
    sheet = workbook.worksheets[sheet_index]

    try:
        new_workbook = load_workbook(output_file_path)
        print("Workbook loaded successfully.")
    except FileNotFoundError:
        new_workbook = Workbook()
        print("New workbook created.")

    new_sheet = new_workbook.active

    print("Begin translating......")
    row_index = 1 
    for row in sheet.iter_rows():
        print(f"> > > begin the translation of row: {row_index}")
        if any(cell.value and isinstance(cell.value, str) and contains_html(cell.value) for cell in row):
            continue 

        for col_index, cell in enumerate(row, start=1):
            if cell.value and isinstance(cell.value, str) and is_french(cell.value):
                # translation
                translated = translator.translate(cell.value, src='fr', dest='en')
                time.sleep(1)
                new_sheet.cell(row=row_index, column=col_index, value=translated.text)
                # print(f"Original: {cell.value} \n Translated: {translated.text} \n")
            else:
                # save
                new_sheet.cell(row=row_index, column=col_index, value=cell.value)

        row_index += 1

    # save
    new_workbook.save(output_file_path)
    print('Successfully saved the translated version without HTML rows.')


if __name__ == "__main__":
    # # parquet_to_excel('./input/train-00000-of-00001.parquet', './input/excel-version.xlsx')
    # # deepl_translate_excel('output.xlsx', 'translated.xlsx', 'EN', 'a948acb3-8e72-41d2-8457-0bd5fc1ddb49:fx')
    # sheet_index = [0, 1, 2] # [0,1,2]
    # for index in sheet_index:
    #     google_translate_excel("./input/excel-version.xlsx", f"./output/{index}-EN-train-00000-of-00001.xlsx", sheet_index=index)
    #     # google_translate_excel("./input/test.xlsx", f"./output/{index}-EN-test.xlsx", sheet_index=index)
    
    # convert_csv_to_xlsx("./input/train1.csv", "./input/train1.xlsx")

    index = 0
    google_translate_excel("./input/train1.xlsx", f"./output/EN-train1.xlsx", sheet_index=index)
