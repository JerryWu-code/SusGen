import time
import pandas as pd
import requests
import argparse
from googletrans import Translator
from openpyxl import load_workbook, Workbook
import re

def parquet_to_excel(parquet_path, excel_path):
    df = pd.read_parquet(parquet_path, engine='pyarrow')
    df.to_excel(excel_path, index=False)

    print(f"Data has been successfully written to {excel_path}")

def deepl_translate_excel(input_excel_path, output_excel_path, target_language, deepl_api_key):
    df = pd.read_excel(input_excel_path)

    # DeepL API URL
    url = "https://api-free.deepl.com/v2/translate"

    for column in df.columns:
        for i in range(len(df)):
            text = df.at[i, column]
            if pd.notna(text):  # None
                params = {
                    'auth_key': deepl_api_key,
                    'text': text,
                    'source_lang': 'KO',
                    'target_lang': target_language
                }
                response = requests.post(url, data=params)
                if response.status_code == 200:
                    translated_text = response.json()['translations'][0]['text']
                    df.at[i, column] = translated_text
                else:
                    print(f"Error translating text: {response.text}")

    # new excel 
    df.to_excel(output_excel_path, index=False)

    print(f"Translation completed and saved to {output_excel_path}")

def is_hangul(text):
    # check korean
    return any('\uac00' <= char <= '\ud7a3' for char in text)

def contains_html(text):
    # check html
    return bool(re.search(r'<[^>]+>', text))

def google_translate_excel(input_file_path, output_file_path, sheet_index):
    translator = Translator(timeout=None)

    workbook = load_workbook(filename=input_file_path)
    sheet = workbook.worksheets[sheet_index]

    try:
        new_workbook = load_workbook(output_file_path)
        print("Workbook loaded successfully.")
    except FileNotFoundError:
        new_workbook = Workbook()
        print("New workbook created.")

    new_sheet = new_workbook.active
    # if f"Sheet{sheet_index+1}" in new_workbook.sheetnames:
    #     new_sheet = new_workbook[f"Sheet{sheet_index+1}"]
    # else:
    #     new_sheet = new_workbook.create_sheet(f"Sheet{sheet_index+1}")

    print("Begin translating......")
    row_index = 1 
    for row in sheet.iter_rows():
        if row_index%10 == 0:
            print(f"> > > begin the translation of row: {row_index}")
        if any(cell.value and isinstance(cell.value, str) and contains_html(cell.value) for cell in row):
            continue 

        for col_index, cell in enumerate(row, start=1):
            if cell.value and isinstance(cell.value, str) and is_hangul(cell.value):
                # translation
                translated = translator.translate(cell.value, src='ko', dest='en')
                time.sleep(1)
                new_sheet.cell(row=row_index, column=col_index, value=translated.text)
                # print(f"Original: {cell.value} \n Translated: {translated.text} \n")
            else:
                # save
                new_sheet.cell(row=row_index, column=col_index, value=cell.value)

        row_index += 1

    # save
    new_workbook.save(output_file_path)
    print('Successfully saved the translated version without HTML rows.')


if __name__ == "__main__":
    # input_file = "./classes/SA-4680.parquet"
    # excel_file = "./classes/SA-4680.xlsx"
    # out_file = "./output/EN-SA-4680.xlsx"
    
    parser = argparse.ArgumentParser(description='Translate data into English.')
    parser.add_argument('--input', type=str, required=True, help='Input parquet data.')
    parser.add_argument('--excel', type=str, required=True, help='Excel data.')
    parser.add_argument('--output', type=str, required=True, help='Output data.')
   
    args = parser.parse_args()
    parquet_to_excel(args.input, args.excel)

    sheet_index = [0] # [0,1,2]
    for index in sheet_index:
        google_translate_excel(args.excel, args.output, sheet_index=index)