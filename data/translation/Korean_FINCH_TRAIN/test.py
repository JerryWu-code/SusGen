from googletrans import Translator
from openpyxl import Workbook, load_workbook
import re
import time

def is_hangul(text):
    # 检查文本中是否包含韩文字符
    return any('\uac00' <= char <= '\ud7a3' for char in text)

def contains_html(text):
    # 使用正则表达式检查文本是否包含HTML标签
    return bool(re.search(r'<[^>]+>', text))

def google_translate_excel(input_file_path, output_file_path, sheet_index):
    # 创建翻译器实例
    translator = Translator()

    # 加载Excel工作簿
    workbook = load_workbook(filename=input_file_path)
    # sheet = workbook.active  # 操作第一个工作表
    sheet = workbook.worksheets[sheet_index]

    # 创建一个新的工作簿来保存翻译后的数据
    new_workbook = Workbook()
    # new_sheet = new_workbook.active
    new_sheet = new_workbook.worksheets[sheet_index]

    # 遍历每个单元格，跳过包含HTML的行
    print("Begin translating......")
    row_index = 1  # 起始行号
    for row in sheet.iter_rows():
        # 检查整行是否包含HTML
        if any(cell.value and isinstance(cell.value, str) and contains_html(cell.value) for cell in row):
            continue  # 如果行包含HTML，则跳过这一行

        for col_index, cell in enumerate(row, start=1):
            if cell.value and isinstance(cell.value, str) and is_hangul(cell.value):
                # 翻译韩文到英文
                translated = translator.translate(cell.value, src='ko', dest='en')
                time.sleep(1)
                new_sheet.cell(row=row_index, column=col_index, value=translated.text)
                print(f"Original: {cell.value} \n Translated: {translated.text} \n")
            else:
                # 将原始数据复制到新表中
                new_sheet.cell(row=row_index, column=col_index, value=cell.value)

        row_index += 1

    # 保存翻译后的工作簿
    new_workbook.save(output_file_path)
    print('Successfully saved the translated version without HTML rows.')


# 使用示例
google_translate_excel('test.xlsx', 'EN-test.xlsx', sheet_index=0)
